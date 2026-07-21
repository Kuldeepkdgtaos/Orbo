import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def build_excel(
    standup_name: str,
    team_name: str,
    date: str,
    rollup_markdown: str,
    key_wins: list[str],
    key_blockers: list[str],
    per_person: list[dict],
    utterances: list[dict],
) -> bytes:
    wb = Workbook()

    # Sheet 1: Rollup
    ws1 = wb.active
    ws1.title = "Rollup"
    title_cell = ws1.cell(row=1, column=1, value=f"Standup Summary — {team_name} — {date}")
    title_cell.font = Font(bold=True, size=14)
    ws1.merge_cells("A1:D1")
    ws1.cell(row=2, column=1, value=standup_name)
    ws1.cell(row=4, column=1, value="Executive Overview").font = Font(bold=True)
    ws1.cell(row=5, column=1, value=rollup_markdown[:500])
    ws1.cell(row=5, column=1).alignment = Alignment(wrap_text=True)
    ws1.row_dimensions[5].height = 60

    ws1.cell(row=7, column=1, value="Key Wins").font = Font(bold=True)
    for i, win in enumerate(key_wins, start=8):
        ws1.cell(row=i, column=1, value=f"- {win}")

    row_offset = 8 + len(key_wins) + 1
    ws1.cell(row=row_offset, column=1, value="Blockers").font = Font(bold=True)
    for i, blocker in enumerate(key_blockers, start=row_offset + 1):
        ws1.cell(row=i, column=1, value=f"- {blocker}")

    ws1.column_dimensions["A"].width = 80

    # Sheet 2: Per Person
    ws2 = wb.create_sheet("Per Person")
    row = 1
    for person in per_person:
        ws2.cell(row=row, column=1, value=person["name"]).font = Font(bold=True, size=12)
        ws2.merge_cells(f"A{row}:C{row}")
        row += 1
        for field in ("yesterday", "today", "blockers"):
            ws2.cell(row=row, column=1, value=field.title()).font = Font(bold=True)
            ws2.cell(row=row, column=2, value=person.get(field, "")).alignment = Alignment(wrap_text=True)
            ws2.merge_cells(f"B{row}:C{row}")
            row += 1
        row += 1

    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 70

    # Sheet 3: Full Transcript
    ws3 = wb.create_sheet("Full Transcript")
    headers = ["Timestamp", "Speaker", "Text"]
    for col, header in enumerate(headers, start=1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = "A1:C1"

    for row_idx, utt in enumerate(utterances, start=2):
        ws3.cell(row=row_idx, column=1, value=utt.get("started_at", ""))
        ws3.cell(row=row_idx, column=2, value=utt.get("speaker_label", ""))
        ws3.cell(row=row_idx, column=3, value=utt.get("text", ""))

    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 25
    ws3.column_dimensions["C"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
